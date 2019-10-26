# Levenshtein Distance
# string a,b; characters are 1-indexed. i.e, a1, a2, a3...; b1, b2, b3
# if min(i,j) = 0, then max(i,j); otherwise min(A,B,C)
# A compares string a with characters up till an-1 with string b
# B compares string a with string b with characters up till bn-1
# C compares string a with characters up till an-1 with string b with characters up till bn-1
# As method C deleted the final string which could mean potential one edit, if the an=bn, then no need edit, otherwise would have one edit
# find i and j by finding the length of the strings, as the length starts from 1
# len(a) = i, len(b)=j
# assume string b is the target string, to match a with b, if A is the min -> deletion as a needs to delet one string;
#if B is the min -> insertation, as a needs to add one string;
# if C is the min -> substitute, as a needs to change certain characters to match b
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#need to do:
#lower case all the names; add more situations to the name possiblities

from functools import reduce #-> for more efficient iterate calculation
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Import NameLibrary for name mactching and appending new names
import sys
import os
scriptpath1 = r"./NameLibrary.py"
sys.path.append(os.path.abspath(scriptpath1))
scriptpath2 = r"./bin.py"
sys.path.append(os.path.abspath(scriptpath2))
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Border
import re
import pandas as pd

with open(scriptpath1, "r+") as f:
    NameDict = json.load(f, strict = False)

with open(scriptpath2,"r+") as f2:
    Bin = json.load(f2,strict=False)

#define levenshtein distance function to be the foundation
def lev(a, b):
    if a == "":
        return len(b) # if a == "", then len(a) -> i = 0, while len(b) -> j; min(0,j) = 0, therefore lev(a,b) = max (0,j) = j

    if b == "":
        return len(a) # if b == "", then len(b) -> j = 0, while len(a) -> i; min(i,0) = 0, therefore lev(a,b) = max(i,0) = i

    if a[-1] == b[-1]:
        cost = 0  # a[-1] = ai, b[-1]=bj, if ai = bj, then deleting both final strings would not result in potential edit

    else:
        cost =1  # a[-1] = ai, b[-1]=bj, if ai <> bj, then deleting both final strings would result in potential one edit
                 # can assign any number as weight -> substitution can be more costy than deletion/insertation

    other = min([lev(a[:-1], b) + 1,  # A: a[:-1] -> string a with characters up till an-1; deleting a character itself has one edit

                 lev(a, b[:-1]) + 1, # B: b[:-1] -> string b with characters up till an-1; deleting a character itself has one edit

                 lev(a[:-1], b[:-1]) + cost])  # C  # if min(i,j) = 0, then lev(a,b) = max (a,b); otherwise lev(a,b)=min(A, B, C)


    #ratio = other/length

    return other
    #return ratio

def length(a,b):
    length = len(a)+len(b)
    return length

def ratio(a,b):
    ratio = (1-round(lev(a,b)/length(a,b),3))*100
    return ratio

def sort(a):
    a = "".join(sorted(a.split()))
    return a
def trim(a):
    a = list(filter(lambda l: l!= '"',[l.strip() for l in a]))
    return a
#-------------------------------------------------------------------------
# read excel file and put column to python list
#file = r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\ftse_100_list.xlsx"

file = r"./ftse100_list.xlsx"
df = pd.read_excel(file, sheet_name=0)
mylist = df['Full Name'].tolist()
namelist = []
list2 = []
list3 = []

for n in mylist:
    match = re.search(r'PLC\s.*', n)
    if match:
        n = n.replace(match.group(),"")
        namelist.append(n)
    else:
        match = re.search(r'ORD\s.*',n)
        if match:
            n = n.replace(match.group(),"")
            namelist.append(n)
namelist = trim(namelist)

for nn in namelist:
    nn = nn.split()
    if nn[-1] in ['CO','AG','LD','GROUP','LTD','INTERNATIONAL','HOLDINGS','HLDGS']:
        nn = " ".join(nn[:-1])
        list2.append(nn)
    elif "".join(nn[-2:]) == "INVTST":
        nn = " ".join(nn[:-2])
        list2.append(nn)
    else:
        nn = " ".join(nn)
        list2.append(nn)

for i in list2:
    match = re.search(r'\([^()]*\)',i)
    if match:
        i = i.replace(match.group(),"")
        list3.append(i)
    else:
        match = re.search(r'GROUP\s.*',i)
        if match:
            i = i.replace(match.group(),"")
            list3.append(i)
        else:
            match = re.search(r'GROUP',i)
            if match:
                i= i.replace(match.group(),"")
                list3.append(i)
            else:
                list3.append(i)
list3 = [l.replace("&", "")for l in list3]
list3 = trim(list3)

wb = openpyxl.load_workbook(file)
s = wb.get_sheet_by_name('Sheet1')
row = 2
for nl in list3:
    s.cell(row, 8).value = nl
    row += 1
head = s.cell(1, 8)
head.value = "Clean Name"
head.font = Font(bold=True)
wb.save(file)

#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#compare with report file
#report = r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx"
report = r'./report.xlsx'
#sheetname = ['Library', 'PSL', 'Draft']
sheetname = ['Library']
for sh in sheetname:
    df = pd.read_excel(report, sheet_name=sh)
    acct = df['accname'].tolist()
    #print(acct)
    matching = list(NameDict.keys())
    ftse = list(NameDict.values())
    wrongname = list(Bin.keys())
    ticker = list(Bin.values())

   # wb2 = openpyxl.load_workbook(
    #    r'\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx')
    wb2 = openpyxl.load_workbook(report)
    Tab = wb2.get_sheet_by_name(sh)
    symbol = []
    existing = []
    checking = []
    checkingname= []

    for acc in acct:
        if acc in matching:
            accindex = acct.index(acc)
            matchingindex = matching.index(acc)
            ftseticker = ftse[matchingindex]
            existing.append(acc)
            Tab.cell(accindex+2,3).value = ftseticker
            wb2.save(report)

        else:
            checking.append(acc)
            checkingname.append(acc)

# name matching for new names

    for n in list3:
        nj = "".join(n.split())

        if n == nj:
            n = r'\b' + n + r'\s'
            for m in checking:
                mindex = checking.index(m)
                cus = checkingname[mindex]
                accindex = acct.index(cus)
                acc = acct[accindex]
                match = re.search(n,m,flags=re.I)
                if match and m not in wrongname:
                    print(m,n)
                    user = input("do you want to add it to the dictionary? y/n: ")
                    if user == "y":
                        matching.append(cus)
                        ftse.append(nj)
                        Tab.cell(mindex+1, 3).value = nj
                        wb2.save(report)
                    #    wb2.save(r'\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx')

                    else:
                        wrongname.append(m)
                        ticker.append(nj)

        else:
            symbol.append(n)
            for s in symbol:
                for m in checking:
                    mindex = checking.index(m)
                    cus = checkingname[mindex]
                    accindex = acct.index(cus)
                    acc = acct[accindex]
                    match = re.search(n,m, flags = re.I)
                    if match and m not in wrongname:
                        print(m,n)
                        user = input("do you want to add it to the dictionary? y/n: ")
                        if user == "y":
                            matching.append(cus)
                            ftse.append(s)
                            Tab.cell(mindex+1, 3).value = s
                            wb2.save(report)
                            #wb2.save(r'\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx')
                        else:
                            wrongname.append(m)
                            ticker.append(nj)


    # above part misses royal shell and lloyds

   # for acctname in acct:
    #    for a, b in NameDict.items():
     #       if acctname == a:
      #          tickersymbol = b
       #         acctindex = acct.index(acctname)+1
        #        if Tab.cell(acctindex+1, 3).value:
         #           pass
          #      else:
           #         Tab.cell(acctindex+1, 3).value = b
            #        wb2.save(report)
                   # wb2.save(r'\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx')


my_dict = dict()
for f, m in zip(ftse, matching):
    my_dict[m] = f
    dict.update(my_dict)

NameDict.update(my_dict)

with open(scriptpath1, 'w') as outfile:
    json.dump(NameDict, outfile)

bin_dict = dict()
for t, w in zip(ticker, wrongname):
    bin_dict[w] = t
    dict.update(bin_dict)

Bin.update(bin_dict)

with open(scriptpath2, "w") as outfile2:
    json.dump(Bin, outfile2)









