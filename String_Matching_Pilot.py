# Levenshtein Distance 
# string a,b; characters are 1-indexed. i.e, a1, a2, a3...; b1, b2, b3
# if min(i,j) = 0, then max(i,j); otherwise min(A,B,C)
# A compares string a with characters up till an-1 with string b  
# B compares string a with string b with characters up till bn-1  
# C compares string a with characters up till an-1 with string b with characters up till bn-1
# As method C deleted the final string which could mean potential one edit, if the an=bn, then no need edit, otherwise would have one edit
# find i and j by finding the length of the strings, as the length starts from 1
# len(a) = i, len(b)=j
# assume string b is the target string, to match a with b, if A is the min -> deletion as a needs to delet one string; 
#if B is the min -> insertation, as a needs to add one string;
# if C is the min -> substitute, as a needs to change certain characters to match b
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#need to do:
#lower case all the names; add more situations to the name possiblities

from functools import reduce #-> for more efficient iterate calculation
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Import NameLibrary for name mactching and appending new names
import sys
import os
scriptpath1 = r".\NameLibrary.py"
sys.path.append(os.path.abspath(scriptpath1))
scriptpath2 = r".\bin.py"
sys.path.append(os.path.abspath(scriptpath2))
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Border
import re 
import pandas as pd 

with open(r".\NameLibrary.py", "r+") as f:
    NameDict = json.load(f)

with open(r".\bin.py","r+") as f2:
    Bin = json.load(f2)

#define levenshtein distance function to be the foundation
def lev(a, b):
    if a == "":
        return len(b) # if a == "", then len(a) -> i = 0, while len(b) -> j; min(0,j) = 0, therefore lev(a,b) = max (0,j) = j

    if b == "":
        return len(a) # if b == "", then len(b) -> j = 0, while len(a) -> i; min(i,0) = 0, therefore lev(a,b) = max(i,0) = i

    if a[-1] == b[-1]:
        cost = 0  # a[-1] = ai, b[-1]=bj, if ai = bj, then deleting both final strings would not result in potential edit

    else:
        cost =1  # a[-1] = ai, b[-1]=bj, if ai <> bj, then deleting both final strings would result in potential one edit
                 # can assign any number as weight -> substitution can be more costy than deletion/insertation

    other = min([lev(a[:-1], b) + 1,  # A: a[:-1] -> string a with characters up till an-1; deleting a character itself has one edit

                 lev(a, b[:-1]) + 1, # B: b[:-1] -> string b with characters up till an-1; deleting a character itself has one edit

                 lev(a[:-1], b[:-1]) + cost])  # C  # if min(i,j) = 0, then lev(a,b) = max (a,b); otherwise lev(a,b)=min(A, B, C)
    
 
    #ratio = other/length
    
    return other
    #return ratio

def length(a,b):
    length = len(a)+len(b)
    return length

def ratio(a,b):
    ratio = (1-round(lev(a,b)/length(a,b),3))*100
    return ratio

def sort(a):
    a = "".join(sorted(a.split()))
    return a


# Name Matching two lists with exact length
foo = "Lexis Nexis"
boo = "Nexis Lexis"

foo_list = sorted(foo.split()) # -> ["Lexis", "Nexis"]
boo_list = sorted(boo.split()) # -> ["Lexis", "Nexis"]
sum = 0
for i,j in zip(range(len(foo_list)),range(len(boo_list))):
    ld=(lev(foo_list[i],boo_list[j]))
    sum=sum+ld
#print(sum)  

# lambda is anonymous function -> def sum_array(accumulator, entry):   return accumulator + entry #->lambda x, y
#sum and sum2 basically the same thing
sum2 = reduce(lambda x,y: x +y , [lev(foo_list[i], boo_list[i]) for i in range(len(foo_list))])
#print(sum2)
# Name Matching essensially two lists with no length restriction
foo_join = "".join(foo_list)
boo_join = "".join(boo_list)
#print(foo_join,boo_join)
#print(lev(foo_join, boo_join))

#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Compare names against a name dictionary, append the dictionary with new names when there's any manual input
#print(NameDict)

# save manually added key/value to a new dictionary and update this dict to the original NewDict
#my_dict = dict()
#user_input = input("please enter the key and value separated by comma: ")
#key, value = user_input.split(",")
#my_dict[key] = value
#dict.update(my_dict)
#print(my_dict)
#NameDict.update(my_dict)

with open(r'.\NameLibrary.py','w') as outfile:
    json.dump(NameDict,outfile)

#print(NameDict)

#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#compare the name against the key from the dic, when ratio is greater than 95, give me the value;otherwise pop up user input with that name as key for inputting value
#f_name = "You"
#keys = list(NameDict.keys())
#values = list(NameDict.values())
#s_name = ""
#for key, value in zip(keys, values):
 #   if ratio(f_name, key) >=85:
        #print(ratio(f_name,key))
  #      s_name = value
        #print(s_name)
   # else:
    #    pass
#if s_name == "":
 #       user_input = input("please enter the surname: ")
  #      key, value = f_name, user_input
   #     NameDict[f_name] = user_input
    #    dict.update(NameDict)
     #   with open(r'.\NameLibrary.py', 'w') as outfile:
      #      json.dump(NameDict, outfile)

# f_name from a list in excel, look up to populate the surname sheet
wb = openpyxl.load_workbook(r'.\test.xlsx')
woo = wb.get_sheet_by_name("Sur")
coo = wb.get_sheet_by_name("Fir")
sur_list = []
fir_list = []
for cell in woo['A']:
    #print(cell.value)
    sur_list.append(cell.value)
    if cell.value is None:
        break
#print(sur_list)

for cell in coo['A']:
   # print(cell.value)
    fir_list.append(cell.value)
    if cell.value is None:
        break


keys = list(NameDict.keys())
values = list(NameDict.values())

for fir in fir_list:
    findex = fir_list.index(fir)
    fir = "".join(sorted(fir.split()))
    for key, value in zip(keys, values):
        if ratio(fir, key) >= 85:
            #print(ratio(fir,key))
            sur = value
            #print(fir,sur)
            if sur in sur_list:
                sindex = sur_list.index(sur)
                #print(sindex)
                fnum = coo.cell(findex+1,2).value
                #print(fir,findex,fnum)
                s = woo.cell(sindex+1, 2).value
                woo.cell(sindex+1, 2).value = fnum
                #print(fir, sur, s)
                wb.save(r"C:\Users\chenyx\Documents\Evelyn\Practise\Python learning\FTSE100\test.xlsx")
            else:
                #print(sur,"not in")  #"AS" not coming up because of surname lowercase
                pass
                # 
#--------------------------------------------------------------------------------------------------------------------------------------------------
# find a word in a long name
str = 'Charity Commission for Northern Ireland'
match = re.search(r'Northern Ireland', str)
if match:
   # print('found', match.group())
    location = match.group()
#    print(location)
    str = str.replace(location, "").strip()
 #   print(str)

# read excel file and put column to python list
file = r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\ftse_100_list.xlsx"
df = pd.read_excel(file, sheet_name=0)
mylist = df['Full Name'].tolist()
namelist = []
list2 = []
list3 = []

for n in mylist:
    match = re.search(r'PLC\s.*', n) 
    if match:
        n = n.replace(match.group(),"")
        namelist.append(n)
    else:
        match = re.search(r'ORD\s.*',n)
        if match:
            n = n.replace(match.group(),"")
            namelist.append(n)
        


for nn in namelist:
    match = re.search(r'\([^()]*\)', nn)
    if match:
     #   print(match.group())
        nn = nn.replace(match.group(), "")
        list2.append(nn)
    else:   
        match = re.search(r'\sCO\s.*',nn)
        if match:
            nn = nn.replace(match.group(),"")
            list2.append(nn)
        else:
            match = re.search(r'AG\s.*',nn)
            if match:
                nn = nn.replace(match.group(),"")
                list2.append(nn)
            else:
                match = re.search(r'LD\s.*', nn)
                if match:
                    nn = nn.replace(match.group(), "")
                    list2.append(nn)
                else:
                    list2.append(nn)



list2 = [l.replace("&", "") for l in list2]
list2 = [l.replace("GROUP", "") for l in list2]
list2 = [l.replace("LTD", "") for l in list2]
list2 = [l.replace("INTERNATIONAL", "") for l in list2]
list2 = [l.replace("HOLDINGS", "") for l in list2]
list2 = [l.replace("HLDGS", "") for l in list2]
list2 = [l.replace("INV TST", "") for l in list2]


list2 = list(filter(lambda l: l != '"', [l.strip() for l in list2]))

#print(list2)


wb2 = openpyxl.load_workbook(file)
s = wb2.get_sheet_by_name('Sheet1')
row = 2
for nl in list2:
    s.cell(row, 8).value = nl
    #wb2.save(file)
    row += 1
head = s.cell(1, 8)
head.value = "Clean Name"
head.font = Font(bold=True)
#wb2.save(file)

#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#compare with report file
report = r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Old.xlsx"
df = pd.read_excel(report, sheet_name='Library')
acct = df['accname'].tolist()
#print(acct)
matching = list(NameDict.keys())
ftse = list(NameDict.values())
wrongname = []
ticker = []
for n in list2:
    nj = "".join(n.split())
    if n == nj:
        n = r'\b' + n + r'\s'
        for m in acct:
            match = re.search(n,m,flags=re.I)
            if match:  
                value = match.group()
                if value in ftse:
                    print(value)
                else:
                    print(n.index,match.group(), m)
                    user = input("type the number you want to add or type N to exit: ")
                    if user != "N":
                        ftse.append(nj)
                        matching.append(m)
                    else:
                        wrongname.append(m)
                        ticker.append(nj)
    else:
        for m in acct:
            match = re.search(n,m, flags = re.I)
            if match:
                #print(match.group(),m)
                ftse.append(n)
                matching.append(m)


my_dict = dict()
for f, m in zip(ftse, matching):
    my_dict[m]=f
    dict.update(my_dict)

NameDict.update(my_dict)

with open(r'.\NameLibrary.py', 'w') as outfile:
    json.dump(NameDict, outfile)

bin_dict = dict()
for t, w in zip(ticker, wrongname):
    bin_dict[w]=t
    dict.update(bin_dict)

Bin.update(bin_dict)

with open(r".\bin.py","w") as outfile2:
    json.dump(Bin,outfile)








 
