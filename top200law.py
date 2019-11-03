#-----import modules--------------------------------------
import pandas as pd
import requests
import urllib.request
import urllib
import os
import time
from bs4 import BeautifulSoup
import re
import sys
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Border
import shutil
import datetime

#-----functions--------------------------------------------
def lev(a, b): # levenshtein distance for string matching
    if a == "":
        return len(b)
    if b == "":
        return len(a)
    if a[-1] == b[-1]:
        cost = 0
    else:
        cost =1
    other = min([lev(a[:-1], b) + 1,
                 lev(a, b[:-1]) + 1,
                 lev(a[:-1], b[:-1]) + cost])
    return other

def length(a,b):
    length = len(a)+len(b)
    return length

def ratio(a,b):
    ratio = (1-round(lev(a,b)/length(a,b),3))*100
    return ratio

def sort(a):
    a = "".join(sorted(a.split()))
    return a
def trim(a):
    a = list(filter(lambda l: l!= '"',[l.strip() for l in a]))
    return a

def join(a):
    a = "".join(a.split())
    return a

def lower(a):
    a = "".join(trim(a)).lower()
    return a

def initial (a,b): # for abbreviation comparison
    a = a.split(' ')
    b = b.split(' ')
    a = list(filter(lambda i: i != '', a))
    b = list(filter(lambda i: i != '', b))
    c = []

    if a[0].isupper() and b[0].isupper():
        inratio = ratio(a[0],b[0])
        return inratio

    elif a[0].isupper() == True and len(a[0])>1 and b[0].isupper() == False:
        if len(b) >=len(a[0]):
            b = b[:len(a[0])]
            for i in b:
                c.append(i[0])
            c = "".join(c)
            inratio = ratio(a[0],c)
            return inratio
        else:
            pass
    elif a[0].isupper()==False and b[0].isupper() == True and len(b[0]) >1:
        if len(a) >= len(b[0]):
            a = a[:len(b[0])]
            for i in a:
                c.append(i[0])
            c = "".join(c)
            inratio = ratio(b[0],c)
            return inratio
        else:
            pass
    else:
        pass

def removeand (a):
    a = a.split(' ')
    a =" ".join(list(filter(lambda i: i != '&' and i != 'and', a)))
    return a

#------get FTSE100 and Top 200 Law lists------------------

# for FTSE 100 index
pages = [1, 2, 3, 4, 5, 6]
output_file = r"./test/ftse100.xlsx"
#output_file = r'\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\ftse_100_list.xlsx'
lst=[]
for page in pages:
    url = r'https://www.londonstockexchange.com/exchange/prices-and-markets/stocks/indices/summary/summary-indices-constituents.html?index=UKX&page={}'.format(
        page)
    response = requests.get(url)
    soup = BeautifulSoup(response.text,"html.parser")
    table = soup.find('table',{'class':'table_dati'})
    table_rows = table.findAll('tr')
    l=[]
    for tr in table_rows:
        td=tr.findAll('td')
        row = list(filter(lambda r: r != '""', [tr.text.strip() for tr in td]))

        url_name = tr.find('a')
        link = r'http://www.londonstockexchange.com' + url_name.attrs['href']
        response = requests.get(link)
        soup = BeautifulSoup(response.text,"html.parser")
        t = soup.find('h1',{'class':'tesummary'})
        name = t.text.strip()

        if name != 'FTSE 100':
            ticker = re.search(r'.*\s', name)
            name = name.replace(ticker.group(),"").strip()
            row.append(name)

        while("" in row):
            row.remove("")

        l.append(row)
    lst.extend(l)
FTSE = [e for e in lst if e!=[]]
df=pd.DataFrame(FTSE,columns=['Code','Name','Cur','Price','+/-','%+/-','Full Name'])
df.to_csv(output_file, index=False, encoding = 'utf-8-sig')
df.to_excel(output_file, index=False)


# for Top 200 law firms
output_file_2 = r"./test/law200.xlsx"
#output_file_2 = r'\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\law_firm_list.xlsx'
lst_2 = []
url = r'https://www.thelawyer.com/top-200-uk-law-firms/'
response = requests.get(url)
soup2 = BeautifulSoup(response.text, "html.parser")
table2 = soup2.find('tbody')
table_rows2 = table2.findAll('tr')
a = []
for tr2 in table_rows2:
    td2 = tr2.findAll('td')
    row2 = [tr2.text.strip() for tr2 in td2]
    while("" in row2):
        row2.remove("")
    a.append(row2)
law = [x for x in a if x != []]
head = law[0]
del law[0]
df2 = pd.DataFrame(law,columns=head)
df2.to_csv(output_file_2, index=False, encoding='utf-8-sig') #avoid ()shown as funny characters
df2.to_excel(output_file_2, index=False)


#-------read excel files and clean company names--------------------------------

#--------prepare FTSE100 list--------------------------------------------------
ftsefile = r"./test/ftse100.xlsx"
file = r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\ftse_100_list.xlsx"
fdf = pd.read_excel(ftsefile, sheet_name=0)
ftselist = fdf['Full Name'].tolist()

#each iteration takes place in the updated ftselist
for company in ftselist:
    cindex = ftselist.index(company)
    match = re.search(r'\([^()]*\)',company)
    if match:
        company = company.replace(match.group(),"").strip()
        ftselist[cindex] = company

for company in ftselist:
    cindex = ftselist.index(company)
    match = re.search(r'PLC\s.*', company)
    if match:
        company = company.replace(match.group(),"").strip()
        ftselist[cindex] = company
    else:
        match = re.search(r'ORD\s.*',company)
        if match:
            company = company.replace(match.group(),"").strip()
            ftselist[cindex] = company

for company in ftselist:
    cindex = ftselist.index(company)
    match = re.search(r'GROUP\s.*',company)
    if match:
        company = company.replace(match.group(),"").strip()
        ftselist[cindex] = company


for company in ftselist:
    cindex = ftselist.index(company)
    company = company.split()
    if company[-1] in ['CO','AG','LD','GROUP','LTD','INTERNATIONAL','HOLDINGS','HLDGS']:
        company = " ".join(company[:-1])
        ftselist[cindex] = company
    elif "".join(company[-2:]) == "INVTST":
        company = " ".join(company[:-2])
        ftselist[cindex] = company

ftselist = [company.replace(" & ", " ").strip() for company in ftselist]

fwb = openpyxl.load_workbook(ftsefile)
fsh = fwb['Sheet1']
row = 2
for company in ftselist:
    fsh.cell(row, 8).value = company
    row += 1
fhead = fsh.cell(1, 8)
fhead.value = "Clean Name"
fhead.font = Font(bold=True)
fwb.save(ftsefile)

#----------prepare Top 200 Law Firms List-----------------------------------
lawfile = r"./test/law200.xlsx"
#file = r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\law_firm_list.xlsx"
ldf = pd.read_excel(lawfile, sheet_name=0)
lawlist = ldf['Firm'].tolist()
lawrank = ldf['Rank 2019'].tolist()


#---------name matching with the Report-----------------------------------------------
#get the ftse/law dictionaries ready
ftsenamepath = r"./ftselibrary.py"
sys.path.append(os.path.abspath(ftsenamepath))

ftsebinpath = r"./ftsebin.py"
sys.path.append(os.path.abspath(ftsebinpath))

with open(ftsenamepath, "r+") as f:
    ftseNameDict = json.load(f, strict = False) # strict to allow space between strings

with open(ftsebinpath,"r+") as f2:
    ftseBinDict = json.load(f2, strict = False)

lawnamepath = r"./lawlibrary.py"
sys.path.append(os.path.abspath(lawnamepath))

lawbinpath = r"./lawbin.py"
sys.path.append(os.path.abspath(lawbinpath))

with open(lawnamepath, "r+") as f3:
    lawNameDict = json.load(f3, strict = False)

with open(lawbinpath, "r+") as f4:
    lawBinDict = json.load(f4, strict = False)

#compare with report file
fmatching = list(ftseNameDict.keys())
fticker = list(ftseNameDict.values())
fwrongname = list(ftseBinDict.keys())
fwrongticker = list(ftseBinDict.values())

lmatching = list(lawNameDict.keys())
lfirm =list(lawNameDict.values())
lwrongname = list(lawBinDict.keys())
lwrongfirm = list(lawBinDict.values())

report = r"./report.xlsx"
#report = r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx"
sheetname = ['Library', 'PSL', 'Draft']

for sh in sheetname:
    df = pd.read_excel(report, sheet_name=sh)
    acctlist = df['accname'].tolist()
    acctname = df['accname'].tolist()
    rwb = openpyxl.load_workbook(report)
#    rwb = openpyxl.load_workbook(
#        r'\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx')
    Tab = rwb[sh]

#for ftse100---------------------------------------------------------
    symbol = []
    fexisting = []
    fchecking = []

    for acc in acctlist: #fill straight away the previously paired companies
        if acc in fmatching:
            accindex = acctlist.index(acc)
            matchingindex = fmatching.index(acc)
            ftseticker = fticker[matchingindex]
            fexisting.append(acc)
            Tab.cell(accindex+2,3).value = ftseticker
            rwb.save(report)

        else:
            fchecking.append(acc)

    for company in ftselist: # for other company names
        com = "".join(company.split())

        if company == com: # check if ftse ticker is just one word
            company = r'\b' + company + r'\s'
            for acc in fchecking:
                accindex = fchecking.index(acc)
                cus = fchecking[accindex]
                cusindex = acctlist.index(cus)
                cusname = acctlist[cusindex]
                match = re.search(company,acc,flags=re.I)
                if match and acc not in fwrongname:
                    print(acc,company)
                    user = input("Do you want to add it to the dictionary? y/n: ")
                    if user == "y":
                        matching.append(cus)
                        ftse.append(com)
                        Tab.cell(cusindex+2, 3).value = com
                        rwb.save(report)
                    #    wb2.save(r'\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx')

                    else:
                        fwrongname.append(acc)
                        fwrongticker.append(com)

        else:
            symbol.append(company)
            for company in symbol:
                for acc in fchecking:
                    accindex = fchecking.index(acc)
                    cus = fchecking[accindex]
                    cusindex = acctlist.index(cus)
                    cusname = acctlist[cusindex]
                    match = re.search(company,acc, flags = re.I)
                    if match and acc not in fwrongname:
                        print(acc,company)
                        user = input("Do you want to add it to the dictionary? y/n: ")
                        if user == "y":
                            matching.append(cus)
                            ftse.append(company)
                            Tab.cell(cusindex+2, 3).value = company
                            rwb.save(report)
                            #wb2.save(r'\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx')
                        else:
                            fwrongname.append(acc)
                            fwrongticker.append(company)

# top200 law firms-----------------------------------------------------------------------------------------

    lexisting = []
    lchecking = []
    lcheckingname = []


    for acc in acctlist:
            if acc in lmatching:
                accindex = acctlist.index(acc)
                matchingindex = lmatching.index(acc)
                lawname = lfirm[matchingindex]
                firmname = " ".join(lawname.split())
                lawindex = lawlist.index(firmname)
                lexisting.append(acc)
                Tab.cell(accindex+2,2).value = firmname
                Tab.cell(accindex+2,4).value = lawrank[lawindex]
                rwb.save(report)
            else:
                lchecking.append(acc)
                lcheckingname.append(acc)

    # clean account name
    for acc in lchecking:
        match = re.search(r'\([^()]*\)',acc)
        accindex = lchecking.index(acc)
        if match:
            acc = acc.replace(match.group(),"")
            lchecking[accindex] = acc
    for acc in lchecking:
        accindex = lchecking.index(acc)
        acct = acc.split()
        if acct[-1] in ['LTD','Ltd','Ltd.','LLP','Limited','AG','AG,','Corp','Corporation','Firm','GmbH-UK','Group','Holding','Holdings','Inc.','Ind','Inc','LIMITED','LLC','Llp','London','Co.','S.A','RLLP','SA','Service','Services','SERVICES','Trust','UK']:
            acc =" ".join(acct[:-1])
            lchecking[accindex] = acc
    for acc in lchecking:
        accindex = lchecking.index(acc)
        if acc =='Anderson\xa0Strathern': #name from the database has weird characters for some reason
            acc = 'Anderson Strathern'
            lchecking[accindex] = acc
    for firm in lawlist:
        ff = firm.split(' ')
        for acc in lchecking:
            accindex = lchecking.index(acc)
            cus = lcheckingname[accindex]
            cusindex = acctname.index(cus)
            cusname = acctname[cusindex]
            firmindex = lawlist.index(firm)
            ff = firm.split(' ')
            if initial(acc,firm) == 100 and cusname not in lwrongname:
                print(acc,firm)
                user = input("Do you want to add it into the dictionary? y/n: ")
                if user == "y":
                    lfirm.append(firm)
                    lmatching.append(cusname)
                    Tab.cell(cusindex+2, 2).value = firm
                    Tab.cell(cusindex+2, 4).value = lawrank[firmindex]
                    rwb.save(report)
                else:
                    lwrongfirm.append(firm)
                    lwrongname.append(cusname)


            elif ff[0].isupper()== acct[0].isupper() ==False:
                match = re.search(removeand(firm),removeand(acc),flags=re.I)
                if match and cusname not in lwrongname:
                    print(acc,firm)
                    user = input("Do you want to add it into the dictionary?y/n: ")
                    if user =="y":
                        lfirm.append(firm)
                        lmatching.append(cusname)
                        Tab.cell(cusindex+2,2).value = firm
                        Tab.cell(cusindex+2,4).value = lawrank[firmindex]
                        rwb.save(report)
                    else:
                        lwrongfirm.append(firm)
                        lwrongname.append(cusname)


my_fdict = dict()
for f, m in zip(fticker, fmatching):
    my_fdict[m] = f
    dict.update(my_fdict)

ftseNameDict.update(my_fdict)

with open(ftsenamepath, 'w') as foutfile:
    json.dump(ftseNameDict, foutfile)

bin_fdict = dict()
for t, w in zip(fwrongticker, fwrongname):
    bin_fdict[w] = t
    dict.update(bin_fdict)

ftseBinDict.update(bin_fdict)

with open(ftsebinpath, "w") as fbinoutfile:
    json.dump(ftseBinDict, fbinoutfile)

my_ldict = dict()
for f, m in zip(lfirm, lmatching):
    my_ldict[m] = f
    dict.update(my_ldict)

lawNameDict.update(my_ldict)

with open(lawnamepath, 'w') as loutfile:
    json.dump(lawNameDict, loutfile)

bin_ldict = dict()
for t, w in zip(lwrongfirm, lwrongname):
    bin_ldict[w] = t
    dict.update(bin_ldict)

lawBinDict.update(bin_ldict)

with open(lawbinpath, "w") as lbinoutfile:
    json.dump(lawBinDict, lbinoutfile)

now = datetime.datetime.now()
yearmonth = now.strftime("%b_%Y")
shutil.copy(report,r"./test/report_%s.xlsx") %yearmonth
