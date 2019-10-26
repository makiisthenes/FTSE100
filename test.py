# Levenshtein Distance
# string a,b; characters are 1-indexed. i.e, a1, a2, a3...; b1, b2, b3
# if min(i,j) = 0, then max(i,j); otherwise min(A,B,C)
# A compares string a with characters up till an-1 with string b
# B compares string a with string b with characters up till bn-1
# C compares string a with characters up till an-1 with string b with characters up till bn-1
# As method C deleted the final string which could mean potential one edit, if the an=bn, then no need edit, otherwise would have one edit
# find i and j by finding the length of the strings, as the length starts from 1
# len(a) = i, len(b)=j
# assume string b is the target string, to match a with b, if A is the min -> deletion as a needs to delet one string;
#if B is the min -> insertation, as a needs to add one string;
# if C is the min -> substitute, as a needs to change certain characters to match b
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#need to do:
#lower case all the names; add more situations to the name possiblities

import pandas as pd
import re
from openpyxl.styles import Font, Color, Border
from openpyxl import Workbook
import openpyxl
import json
from functools import reduce  # -> for more efficient iterate calculation
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Import NameLibrary for name mactching and appending new names
import sys
import os
#scriptpath1 = r".\Lawlibrary.py"
scriptpath1 =  r"./Lawlibrary.py"
sys.path.append(os.path.abspath(scriptpath1))
#scriptpath2 = r".\Lawbin.py"
scriptpath2 = r"./Lawbin.py"
sys.path.append(os.path.abspath(scriptpath2))

#with open(r".\Lawlibrary.py", "r+") as f:
with open (scriptpath1,"r+") as f:
    NameDict = json.load(f)

#with open(r".\Lawbin.py", "r+") as f2:
with open (scriptpath2,"r+") as f2:
    Bin = json.load(f2)

#define levenshtein distance function to be the foundation


def lev(a, b):
    if a == "":
        # if a == "", then len(a) -> i = 0, while len(b) -> j; min(0,j) = 0, therefore lev(a,b) = max (0,j) = j
        return len(b)

    if b == "":
        # if b == "", then len(b) -> j = 0, while len(a) -> i; min(i,0) = 0, therefore lev(a,b) = max(i,0) = i
        return len(a)

    if a[-1] == b[-1]:
        # a[-1] = ai, b[-1]=bj, if ai = bj, then deleting both final strings would not result in potential edit
        cost = 0

    else:
        # a[-1] = ai, b[-1]=bj, if ai <> bj, then deleting both final strings would result in potential one edit
        cost = 1
        # can assign any number as weight -> substitution can be more costy than deletion/insertation

    other = min([lev(a[:-1], b) + 1,  # A: a[:-1] -> string a with characters up till an-1; deleting a character itself has one edit

                 # B: b[:-1] -> string b with characters up till an-1; deleting a character itself has one edit
                 lev(a, b[:-1]) + 1,

                 lev(a[:-1], b[:-1]) + cost])  # C  # if min(i,j) = 0, then lev(a,b) = max (a,b); otherwise lev(a,b)=min(A, B, C)

    #ratio = other/length

    return other
    #return ratio


def length(a, b):
    length = len(a)+len(b)
    return length


def ratio(a, b):
    ratio = (1-round(lev(a, b)/length(a, b), 3))*100
    return ratio


def sort(a):
    a = "".join(sorted(a.split()))
    return a

def initial (a,b):
    a = a.split(' ')
    b = b.split(' ')
    a = list(filter(lambda i: i != '', a))
    b = list(filter(lambda i: i != '', b))
    c = []

    if a[0].isupper() and b[0].isupper():
        inratio = ratio(a[0],b[0])
        return inratio

    elif a[0].isupper() == True and len(a[0])>1 and b[0].isupper() == False:
        if len(b) >=len(a[0]):
            b = b[:len(a[0])]
            for i in b:
                c.append(i[0])
            c = "".join(c)
            inratio = ratio(a[0],c)
            return inratio
        else:
            pass
    elif a[0].isupper()==False and b[0].isupper() == True and len(b[0]) >1:
        if len(a) >= len(b[0]):
            a = a[:len(b[0])]
            for i in a:
                c.append(i[0])
            c = "".join(c)
            inratio = ratio(b[0],c)
            return inratio
        else:
            pass
    else:
        pass

def removeand (a):
    a = a.split(' ')
    a =" ".join(list(filter(lambda i: i != '&' and i != 'and', a)))

    return a

#----------------------------------------------------------------------------------------------------------------------------------------
# read excel file and put column to python list
#file = r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\law_firm_list.xlsx"
file = r"./lawfirm_list.xlsx"
df = pd.read_excel(file, sheet_name=0)
namelist = df['Firm'].tolist()

# clean the * character in the end of the names

for n in namelist:
    nindex = namelist.index(n)
    n = re.sub(r'\*',"",n)
    namelist[nindex] = n

# put cleaned name in a new column
wb = openpyxl.load_workbook(file)
s = wb.get_sheet_by_name('Sheet1')
row = 2
for n in namelist:
    s.cell(row, 5).value = n
    wb.save(file)
    row += 1
head = s.cell(1, 5)
head.value = "Clean Name"
head.font = Font(bold=True)
#wb.save(file)


#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#compare with report file
#report = r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx"
report = r'./report.xlsx'
#sheetname = ['Library','PSL','Draft']
sheetname = ['Library']
for sh in sheetname:

    df = pd.read_excel(report, sheet_name= sh)

    acct = df['accname'].tolist()
    acctname = df['accname'].tolist()

    matching = list(NameDict.keys())
    #matching = ['Berrymans Lace Mawer LLP','Holman Fenwick Willan LLP','Reynolds Porter Chamberlain LLP','JMW Solicitors LLP','McMillan Williams Solicitors Limited','AG Service Company Limited','BDB Pitmans LLP','Clyde & Co LLP','Bates Wells & Braithwaite (Ipswich)','Knights Professional Services Ltd.','HCR Legal LLP','Thrings LLP','Anderson Strathern','Sacker & Partner','Bates Wells & Braithwaite Limited','Slaughter & May','Trowers & Hamlins (services) Limited','Mayo Wynne Baxter LLP','Drydens Fairfax']
#    print(matching)
    law = list(NameDict.values())
    wrongname = list(Bin.keys())
    firm = list(Bin.values())

#    wb2 = openpyxl.load_workbook(
#        r"\\Galileo\Public\Legal Intelligence\Customer Segmentation\BA\Ad Hoc Reports & Requests\2019\201909 - September\DAI-2093 - Kenneth Ume - Market Product Penetration Data Request - REPORT\8. WORKINGS_Sep - Copy.xlsx")

    wb2 = openpyxl.load_workbook(report)

    Tab = wb2.get_sheet_by_name(sh)
# clean account name
    for m in acct:
        match = re.search(r'\([^()]*\)',m)
        mindex = acct.index(m)
        if match:
            m = m.replace(match.group(),"")
            acct[mindex] = m
    for m in acct:
        mindex = acct.index(m)
        mm = m.split()
        if mm[-1] in ['LTD','Ltd','Ltd.','LLP','Limited','AG','AG,','Corp','Corporation','Firm','GmbH-UK','Group','Holding','Holdings','Inc.','Ind','Inc','LIMITED','LLC','Llp','London','Co.','S.A','RLLP','SA','Service','Services','SERVICES','Trust','UK']:
            m =" ".join(mm[:-1])
            acct[mindex] = m
    for n in namelist:
        nn = n.split(' ')
        for m in acct:
            mm = m.split(' ')
            if initial(m,n) == 100 and acctname[acct.index(m)] not in matching:
                print(m,n)
                user = input("do you want to add in the library? y/n")
                if user == 'y':
                    matching.append(acctname[acct.index(m)])
                else:
                    wrongname.append(acctname[acct.index(m)])

            elif nn[0].isupper() == mm[0].isupper() == False:
                match = re.search(removeand(n),removeand(m),flags=re.I)
                if match and acctname[acct.index(m)] not in matching:
                    print(m,n)
                    user = input("do you want to add in the library? y/n")
                    if user == 'y':
                        matching.append(acctname[acct.index(m)])
                    else:
                        wrongname.append(acctname[acct.index(m)])



