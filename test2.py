import datetime
import shutil
import openpyxl
import pandas as pd
from openpyxl import load_workbook
#now = datetime.datetime.now()
#yearmonth = now.strftime("%b_%Y")
#print(yearmonth)
#path = r"./test/report_{}.xlsx".format(yearmonth)
#shutil.copy(r"./report.xlsx",path)
#print(path)
#report = r"./test/ftse100.xlsx"
#wb = load_workbook(report)
#wb.create_sheet("sheet1")
#wb.save(report)

report = r"./report.xlsx"
workbook=load_workbook(report,data_only = True)
previous = workbook.get_sheet_by_name('previous')
current = workbook.get_sheet_by_name('Top200LawFirms')

rows=previous[2:previous.max_row]
#print(rows)

df = pd.read_excel(report,sheet_name='Top200LawFirms')
firmnames = df['Name'].tolist()
df2 = pd.read_excel(report,sheet_name='PSL')
finalfirm = df2['Top200Law'].tolist()
#for name in firmnames:
#    index = firmnames.index(name)
#    namerow = index+2
#    if name in finalfirm:
#        current.cell(namerow,5).value = 'Y'
#        workbook.save(report)
#    else:
#        pass
#for row in rows:
#    for cell in row:
#        cell.value= None
#        rownumber = cell.row
#        columnnumber = cell.column
#        currentvalue = current.cell(rownumber,columnnumber).value
#        cell.value=currentvalue
#        workbook.save(report)
#        print(cell.value)
#        print (cell.row,cell.column)

currentPSL = df['PSL']
numPSL = len(list(filter(lambda n: n == 'Y', currentPSL)))
print(numPSL)
